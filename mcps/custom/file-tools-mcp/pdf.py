"""PDF read/write/edit and document conversion handlers for the file-tools MCP.

Enhanced reader with document properties, structure, and security info.
edit_pdf: 19 operations (merge, split, rotate, watermark, replace_text, redact,
annotate, encrypt, etc.)
pdf_to_images: render pages as PNG.
images_to_pdf: combine images into PDF.
write_pdf: HTML/Markdown → PDF via WeasyPrint.
convert_document: LibreOffice headless format conversion.
"""

import asyncio
import contextlib
import os
import re
import shutil
from pathlib import Path

from isolation import run_parse
from shared import (
    _checked_resolved,
    _dropped_note,
    _libreoffice_convert,
    _normalize_operations,
    _op_type,
    _push_image_preview,
    _push_preview,
    _resolve_or_mark,
    _resolve_path,
    _to_agents_relative,
    _WORKER_TMP_SUFFIX,
    logger,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_pages(pages_str, total: int) -> list[int]:
    """Parse page specification to list of 0-based indices.

    Accepts: "all", single int, "1-5", "1-5,8,10-12", or list of ints.
    Input page numbers are 1-based, output is 0-based.
    """
    if pages_str is None or pages_str == "all":
        return list(range(total))
    if isinstance(pages_str, int):
        return [pages_str] if pages_str < total else []
    if isinstance(pages_str, list):
        return [int(p) for p in pages_str if 0 <= int(p) < total]

    result = []
    for part in str(pages_str).split(","):
        part = part.strip()
        # Negative indexing: -1 = last page, -2 = second to last, etc.
        if part.startswith("-") and part[1:].isdigit():
            idx = total + int(part)
            if 0 <= idx < total:
                result.append(idx)
        elif "-" in part:
            a, b = part.split("-", 1)
            start = max(0, int(a) - 1)
            end = min(total, int(b))
            result.extend(range(start, end))
        elif part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < total:
                result.append(idx)
    return result


def _ocr_raster_dpi(page, requested) -> int:
    """Raster resolution for the OCR page swap: never upsample past the scan.

    The swap re-rasterizes the page, so dpi drives output size — a 300 dpi
    re-raster of a 150 dpi scan is several times the bytes for zero OCR gain
    (an 84-page scan once ballooned to 536MB and stalled the host on disk
    I/O). Cap the requested dpi (default 300) at the source image's own
    resolution, inferred from the widest embedded image; text-only pages
    keep the requested value.
    """
    cap = int(requested) if requested else 300
    try:
        images = page.get_images(full=True)
        if images:
            width_px = max(img[2] for img in images)
            src_dpi = width_px * 72.0 / max(page.rect.width, 1.0)
            cap = min(cap, max(72, round(src_dpi)))
    except Exception:
        pass
    return max(72, min(cap, 600))


def _text_style_at(page, rect):
    """Dominant text style under `rect` → (fontsize, color, base14, baseline_y).

    Original embedded fonts are usually subset-encoded and can't be reused
    for new text, so replacements map to the closest Base-14 family from the
    span flags (serif/mono/bold/italic). Falls back to rect-derived metrics
    when no span overlaps (e.g. text inside a Form XObject).
    """
    import fitz

    best, best_overlap = None, 0.0
    for block in page.get_text("dict", clip=rect).get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                sr = fitz.Rect(span["bbox"])
                if not sr.intersects(rect):
                    continue
                overlap = abs((sr & rect).get_area())
                if overlap > best_overlap:
                    best, best_overlap = span, overlap

    if best is None:
        size = max(6.0, rect.height * 0.8)
        return size, (0, 0, 0), "helv", rect.y1 - rect.height * 0.22

    flags = best.get("flags", 0)
    italic, serif = bool(flags & 2), bool(flags & 4)
    mono, bold = bool(flags & 8), bool(flags & 16)
    if mono:
        base = "cobi" if bold and italic else "cobo" if bold else "coit" if italic else "cour"
    elif serif:
        base = "tibi" if bold and italic else "tibo" if bold else "tiit" if italic else "tiro"
    else:
        base = "hebi" if bold and italic else "hebo" if bold else "heit" if italic else "helv"

    c = best.get("color", 0)
    color = ((c >> 16 & 255) / 255, (c >> 8 & 255) / 255, (c & 255) / 255)
    origin = best.get("origin")
    baseline_y = float(origin[1]) if origin else rect.y1 - rect.height * 0.22
    return float(best.get("size", rect.height * 0.8)), color, base, baseline_y


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------


# Stats sweep ceiling per read — word/image/annotation counts on a huge PDF
# are not worth a page-by-page pass over thousands of pages.
_STATS_SAMPLE_PAGES = 500


def read_pdf(path: str, pages: str | None) -> str:
    import fitz

    doc = fitz.open(path)
    meta = doc.metadata or {}
    total = doc.page_count
    result = [f"**PDF**: {Path(path).name} — {total} page(s)"]

    # Document properties
    props = []
    if meta.get("title"):
        props.append(f"Title: {meta['title']}")
    if meta.get("author"):
        props.append(f"Author: {meta['author']}")
    if meta.get("subject"):
        props.append(f"Subject: {meta['subject']}")
    if meta.get("creator"):
        props.append(f"Creator: {meta['creator']}")
    if meta.get("creationDate"):
        props.append(f"Created: {meta['creationDate'][:10]}")
    if props:
        result.append("**Properties**: " + " | ".join(props))

    # Page info
    if total > 0:
        p0 = doc[0]
        w_in = round(p0.rect.width / 72, 2)
        h_in = round(p0.rect.height / 72, 2)
        orient = "landscape" if w_in > h_in else "portrait"
        result.append(f"**Pages**: {total} pages, {w_in}x{h_in}in ({orient})")

    # File size
    file_size = os.path.getsize(path)
    if file_size > 1_000_000:
        result.append(f"**Size**: {file_size / 1_000_000:.1f} MB")
    else:
        result.append(f"**Size**: {file_size / 1_000:.0f} KB")

    # Read range — parsed BEFORE the stats pass so both halves of one read
    # honor the identical range ("N" / "A-B" semantics).
    start, end = 0, total
    if pages:
        parts = pages.split("-")
        start = max(0, int(parts[0]) - 1)
        end = min(total, int(parts[-1]))

    # Content stats, scanned over the read range only and sampled past a cap:
    # a full-document sweep of a huge PDF costs a page-by-page get_text of
    # everything the caller deliberately did NOT ask for.
    stats_end = min(end, start + _STATS_SAMPLE_PAGES)
    stats = []
    word_count = 0
    img_count = 0
    annot_count = 0
    form_fields = 0
    for i in range(start, stats_end):
        page = doc[i]
        text = page.get_text()
        word_count += len(text.split())
        img_count += len(page.get_images(full=False))
        annot_count += len(list(page.annots() or []))
        form_fields += len(list(page.widgets() or []))
    stats.append(f"~{word_count:,} words")
    if img_count:
        stats.append(f"{img_count} images")
    if annot_count:
        stats.append(f"{annot_count} annotations")
    if form_fields:
        stats.append(f"{form_fields} form fields")
    else:
        stats.append("No forms")
    label = "**Content**"
    if pages:
        label += f" (pages {start + 1}–{end})"
    if stats_end < end:
        label += f" (sampled from the first {stats_end - start} pages of the range)"
    result.append(f"{label}: {' | '.join(stats)}")

    # TOC / bookmarks
    toc = doc.get_toc()
    if toc:
        result.append(f"**TOC**: {len(toc)} bookmarks")
        for level, title, page_num in toc[:10]:
            result.append(f"  {'  ' * (level - 1)}{title} (p.{page_num})")
        if len(toc) > 10:
            result.append(f"  ... and {len(toc) - 10} more")

    # Security
    if doc.is_encrypted:
        result.append("**Security**: Encrypted")
    else:
        result.append("**Security**: Not encrypted")

    result.append("")

    # Text extraction. Pages without a text layer (scans) are surfaced as
    # collapsed range markers so mixed documents stay readable, capped by a
    # guidance line instead of silence — the agent can SEE scanned pages via
    # screenshot_document, which OCR can never match on handwriting.
    scanned_run: list[int] | None = None
    scanned_any = False

    def _flush_run() -> None:
        nonlocal scanned_run
        if scanned_run is None:
            return
        a, b = scanned_run
        if a == b:
            result.append(f"--- Page {a + 1}: scanned image, no text layer ---")
        else:
            result.append(f"--- Pages {a + 1}-{b + 1}: scanned images, no text layer ---")
        result.append("")
        scanned_run = None

    for i in range(start, end):
        text = doc[i].get_text().strip()
        if text:
            _flush_run()
            result.append(f"--- Page {i + 1} ---")
            result.append(text)
            result.append("")
        else:
            scanned_any = True
            if scanned_run is None:
                scanned_run = [i, i]
            else:
                scanned_run[1] = i
    _flush_run()
    if scanned_any:
        result.append(
            "(Scanned pages have no extractable text. To READ them, call "
            "screenshot_document on this file in batches of up to 10 pages "
            "(dpi 300 for handwriting or fine print) and read the page images "
            "directly — that also works for handwriting, which OCR cannot "
            "handle. To make the file searchable, use edit_pdf with an ocr "
            "operation and the document's language(s), e.g. language "
            "'ell+eng' for Greek — OCR rewrites the pages, so large documents "
            "take minutes and produce large output files.)"
        )

    doc.close()
    return "\n".join(result)


# ---------------------------------------------------------------------------
# Write PDF (HTML/Markdown → PDF via WeasyPrint)
# ---------------------------------------------------------------------------

# Math delimiters. \( \) / \[ \] / $$ $$ are always math. Single-$ spans only
# count under pandoc's adjacency rules — opening $ immediately followed by
# non-space, closing $ immediately preceded by non-space and not followed by
# a digit — so "$100 and $200" or "$5-$10" never parse as math.
_MATH_PATTERNS = [
    (re.compile(r"\\\[(.+?)\\\]", re.DOTALL), True),
    (re.compile(r"\$\$(.+?)\$\$", re.DOTALL), True),
    (re.compile(r"\\\((.+?)\\\)", re.DOTALL), False),
]
_SINGLE_DOLLAR_RE = re.compile(
    r"(?<![\\$])\$(?![\s$])((?:[^$\n\\]|\\[^\n])+?)(?<![\s\\])\$(?![\d$])"
)
# Segments math must never touch: fenced code blocks and backtick spans
# (markdown), <pre>/<code> blocks (raw HTML input).
_MD_CODE_RE = re.compile(r"(```.*?```|~~~.*?~~~|`[^`\n]*`)", re.DOTALL)
_HTML_CODE_RE = re.compile(r"(<pre\b.*?</pre>|<code\b.*?</code>)", re.DOTALL | re.IGNORECASE)


def _extract_math_spans(content: str, *, is_markdown: bool):
    """Swap math spans for placeholder tokens that survive markdown conversion.

    Returns (content_with_placeholders, spans) where each span is
    (token, latex, display, original_text). Rendering happens after the HTML
    pipeline so LaTeX backslashes never meet the markdown parser.
    """
    import uuid

    spans: list[tuple[str, str, bool, str]] = []

    def _sub_segment(segment: str) -> str:
        def make_repl(display: bool):
            def repl(m):
                token = f"OTOMATH{uuid.uuid4().hex[:12]}X"
                spans.append((token, m.group(1).strip(), display, m.group(0)))
                return token
            return repl

        for pattern, display in _MATH_PATTERNS:
            segment = pattern.sub(make_repl(display), segment)
        if is_markdown:
            segment = _SINGLE_DOLLAR_RE.sub(make_repl(False), segment)
        return segment

    protect = _MD_CODE_RE if is_markdown else _HTML_CODE_RE
    parts = protect.split(content)
    # protect.split alternates non-code / code segments; transform only non-code
    return "".join(
        _sub_segment(part) if i % 2 == 0 else part
        for i, part in enumerate(parts)
    ), spans


def _render_math_spans(html_body: str, spans, errors: list[str]) -> str:
    """Replace placeholder tokens with inline SVG (ziamath). A failed equation
    restores its original text and reports — one bad formula must not sink
    the whole document."""
    from equations import EquationError, latex_to_svg

    for token, latex, display, original in spans:
        try:
            svg, yofst = latex_to_svg(latex, display=display, size=18 if display else 16)
            if display:
                rendered = (
                    '<span style="display:block;text-align:center;margin:0.8em 0">'
                    f"{svg}</span>"
                )
            else:
                rendered = svg.replace(
                    "<svg ", f'<svg style="vertical-align:{yofst:.2f}px" ', 1
                )
        except EquationError as exc:
            errors.append(str(exc))
            rendered = original
        html_body = html_body.replace(token, rendered)
    return html_body


_WRITE_PDF_ADVICE = (
    "Write shorter content per call (split long documents into sections) "
    "or embed smaller images"
)


def _render_pdf_core(full_html: str, path: str) -> None:
    """Worker core: WeasyPrint layout/render — the memory hazard — plus an
    atomic save (a killed child must never truncate an existing PDF)."""
    from weasyprint import HTML

    tmp = path + _WORKER_TMP_SUFFIX
    HTML(string=full_html).write_pdf(tmp)
    os.replace(tmp, path)


async def handle_write_pdf(args: dict) -> str:
    import markdown as md

    path = await _resolve_path(args["path"], writing=True)
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    content = args.get("content", "")
    content_type = args.get("content_type", "markdown")
    custom_css = args.get("css", "")
    page_size = args.get("page_size", "A4")
    margins = args.get("margins", {})

    content, math_spans = _extract_math_spans(
        content, is_markdown=content_type == "markdown"
    )

    if content_type == "markdown":
        html_body = md.markdown(
            content, extensions=["tables", "fenced_code", "nl2br"]
        )
    else:
        html_body = content

    # Pre-resolve img srcs — a re.sub callback can't await.
    src_map: dict[str, str] = {}
    for src in set(re.findall(r'src="([^"]+)"', html_body)):
        with contextlib.suppress(Exception):
            src_map[src] = await _resolve_path(src)

    def resolve_img_src(match):
        resolved = src_map.get(match.group(1))
        return f'src="file://{resolved}"' if resolved else match.group(0)

    html_body = re.sub(r'src="([^"]+)"', resolve_img_src, html_body)

    # After the src= rewriter — injected SVG markup must never meet it
    math_errors: list[str] = []
    html_body = _render_math_spans(html_body, math_spans, math_errors)

    m_top = margins.get("top", "2cm")
    m_bottom = margins.get("bottom", "2cm")
    m_left = margins.get("left", "2cm")
    m_right = margins.get("right", "2cm")

    full_html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
@page {{ size: {page_size}; margin: {m_top} {m_right} {m_bottom} {m_left}; }}
body {{ font-family: 'Liberation Sans', 'DejaVu Sans', sans-serif; font-size: 12pt; line-height: 1.5; color: #333; }}
h1 {{ font-size: 24pt; margin-top: 0; }}
h2 {{ font-size: 18pt; }}
h3 {{ font-size: 14pt; }}
table {{ border-collapse: collapse; width: 100%; margin: 1em 0; }}
th, td {{ border: 1px solid #ccc; padding: 6px 10px; text-align: left; }}
th {{ background: #f0f0f0; font-weight: bold; }}
code {{ background: #f5f5f5; padding: 2px 4px; border-radius: 3px; font-size: 10pt; }}
pre {{ background: #f5f5f5; padding: 12px; border-radius: 5px; overflow-x: auto; }}
img {{ max-width: 100%; }}
{custom_css}
</style></head><body>{html_body}</body></html>"""

    # Assembly above is cheap and needs the async resolver; the WeasyPrint
    # render is the memory hazard and runs in a bounded worker child.
    try:
        await run_parse(_render_pdf_core, full_html, path, _advice=_WRITE_PDF_ADVICE)
    except Exception:
        with contextlib.suppress(OSError):
            os.unlink(path + _WORKER_TMP_SUFFIX)
        raise
    await _push_preview(path)
    msg = f"PDF created: {_to_agents_relative(path)}"
    rendered_eqs = len(math_spans) - len(math_errors)
    if rendered_eqs:
        msg += f" ({rendered_eqs} equation(s) rendered)"
    if math_errors:
        msg += (
            f"\n\nEquation warnings ({len(math_errors)}) — these spans were left "
            "as plain text:\n" + "\n".join(f"  - {e}" for e in math_errors)
        )
    return msg


# ---------------------------------------------------------------------------
# Edit PDF (19 operations)
# ---------------------------------------------------------------------------


_EDIT_PDF_ADVICE = (
    "Split the operation list into smaller edit_pdf calls or work on a "
    "smaller file"
)


async def _preresolve_pdf_ops(ops: list) -> None:
    """Pre-resolve every path-bearing op field in the parent — the proxy
    resolve API is async/HTTP and can never run inside a worker core.
    Failures become marker strings (`_checked_resolved` re-raises them at the
    op's use site so per-op containment matches the old inline awaits)."""
    for op in ops:
        ot = _op_type(op)
        if ot == "merge":
            op["files"] = [await _resolve_or_mark(f) for f in op.get("files", [])]
        elif ot == "split":
            if op.get("output_path"):
                op["output_path"] = await _resolve_or_mark(
                    op["output_path"], writing=True
                )
        elif ot == "add_image":
            src = op.get("image_path") or op.get("path") or op.get("image", "")
            op.pop("path", None)
            op.pop("image", None)
            op["image_path"] = await _resolve_or_mark(src)
        elif ot == "extract_images":
            od = op.get("output_dir")
            # Absolute-only resolve quirk preserved: relative dirs stay raw,
            # and the missing-dir default derives from the (already resolved)
            # main path inside the core.
            if od and str(od).startswith("/"):
                op["output_dir"] = await _resolve_or_mark(od, writing=True)
        elif ot == "ocr":
            if op.get("output_path"):
                op["output_path"] = await _resolve_or_mark(
                    op["output_path"], writing=True
                )


async def handle_edit_pdf(args: dict) -> str:
    # edit_pdf ALWAYS rewrites the source in place at save time (even
    # extraction-only op sets go through the save→os.replace tail), so the
    # input is a write target — resolve it as one so the proxy's write-RBAC
    # fires instead of the read check.
    path = await _resolve_path(args["path"], writing=True)
    if not Path(path).exists():
        return f"Error: File not found: {args['path']}"

    # Normalize ONCE here (it handles double-JSON-encoded op lists) and
    # detect ocr via _op_type over the NORMALIZED ops — scanning the raw
    # arguments would miss aliased (`op`/`operation`/`action`) or
    # string-encoded ocr ops and ship them into the worker.
    ops, dropped = _normalize_operations(args.get("operations"))
    await _preresolve_pdf_ops(ops)

    if any(_op_type(op) == "ocr" for op in ops):
        # OCR stays out of the worker pool ON PURPOSE: the documented
        # multi-minute runs would hog the single parse permit on a 2g
        # sidecar, starving every read_document in the install, and per-page
        # allocations are already bounded (_ocr_raster_dpi caps the raster
        # at the scan's own resolution). It runs in a thread, not on the
        # event loop — pymupdf/tesseract hold the CPU for minutes and would
        # otherwise stall /health and every other session.
        msg = await asyncio.to_thread(_edit_pdf_core, path, ops, dropped)
    else:
        try:
            msg = await run_parse(
                _edit_pdf_core, path, ops, dropped, _advice=_EDIT_PDF_ADVICE
            )
        except Exception:
            with contextlib.suppress(OSError):
                os.unlink(path + _WORKER_TMP_SUFFIX)
            raise
    await _push_preview(path)
    return msg


def _edit_pdf_core(path: str, ops: list, dropped: int) -> str:
    """Worker core: the whole op loop + atomic save. Pure sync — paths were
    pre-resolved by the parent (markers re-raise at the op's use site), no
    session state, no HTTP."""
    import fitz

    doc = fitz.open(path)
    errors = []
    notes = []

    for idx, op in enumerate(ops):
        ot = _op_type(op)
        try:
            # =============================================================
            # PAGE OPERATIONS
            # =============================================================

            if ot == "merge":
                files = op.get("files", [])
                position = op.get("position", "end")
                for f in files:
                    src_path = _checked_resolved(f)
                    if not Path(src_path).exists():
                        errors.append(f"Op #{idx} merge: file not found: {f}")
                        continue
                    src = fitz.open(src_path)
                    if position == "start":
                        doc.insert_pdf(src, to_page=-1, start_at=0)
                    elif isinstance(position, int):
                        doc.insert_pdf(src, start_at=position)
                    else:
                        doc.insert_pdf(src)
                    src.close()

            elif ot == "split":
                page_spec = op.get("pages", "all")
                output_path = op.get("output_path")
                if not output_path:
                    errors.append(f"Op #{idx} split: output_path required")
                    continue
                out = _checked_resolved(output_path)
                Path(out).parent.mkdir(parents=True, exist_ok=True)
                page_indices = _parse_pages(page_spec, doc.page_count)
                new_doc = fitz.open()
                for pi in page_indices:
                    new_doc.insert_pdf(doc, from_page=pi, to_page=pi)
                # Atomic: `out` may name an existing file, and a killed
                # child must never leave it truncated.
                new_doc.save(out + _WORKER_TMP_SUFFIX)
                new_doc.close()
                os.replace(out + _WORKER_TMP_SUFFIX, out)

            elif ot == "rotate_page":
                page_spec = op.get("pages", "all")
                degrees = int(op.get("degrees", 90))
                for pi in _parse_pages(page_spec, doc.page_count):
                    doc[pi].set_rotation(degrees)

            elif ot == "delete_page":
                page_spec = op.get("pages", [])
                indices = _parse_pages(page_spec, doc.page_count)
                # Delete in reverse order to preserve indices
                for pi in sorted(indices, reverse=True):
                    doc.delete_page(pi)

            elif ot == "reorder_pages":
                order = op.get("order", [])
                if order:
                    doc.select([int(i) for i in order])

            elif ot == "insert_page":
                position = int(op.get("position", -1))
                width = float(op.get("width", 595))
                height = float(op.get("height", 842))
                doc.insert_page(position, width=width, height=height)

            elif ot == "crop_page":
                page_spec = op.get("pages", "all")
                rect = op.get("rect")
                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    if rect == "auto":
                        # Auto-detect content boundaries
                        blocks = page.get_text("blocks")
                        if blocks:
                            x0 = min(b[0] for b in blocks) - 10
                            y0 = min(b[1] for b in blocks) - 10
                            x1 = max(b[2] for b in blocks) + 10
                            y1 = max(b[3] for b in blocks) + 10
                            page.set_cropbox(fitz.Rect(x0, y0, x1, y1))
                    elif rect and len(rect) == 4:
                        page.set_cropbox(fitz.Rect(*rect))

            # =============================================================
            # CONTENT OPERATIONS
            # =============================================================

            elif ot == "add_text":
                pi = int(op.get("page", 0))
                if pi >= doc.page_count:
                    errors.append(f"Op #{idx} add_text: page {pi} out of range")
                    continue
                page = doc[pi]
                text = op.get("text", "")
                x = float(op.get("x", 72))
                y = float(op.get("y", 72))
                font_size = float(op.get("font_size", 12))
                color = op.get("color", [0, 0, 0])
                if isinstance(color, list) and len(color) == 3:
                    color = tuple(float(c) for c in color)
                else:
                    color = (0, 0, 0)
                fontname = op.get("font", "helv")
                page.insert_text(
                    fitz.Point(x, y), text,
                    fontsize=font_size, fontname=fontname, color=color,
                )

            elif ot == "add_image":
                pi = int(op.get("page", 0))
                if pi >= doc.page_count:
                    errors.append(f"Op #{idx} add_image: page {pi} out of range")
                    continue
                page = doc[pi]
                img_path = _checked_resolved(op.get("image_path", ""))
                rect = op.get("rect", [50, 50, 200, 200])
                page.insert_image(fitz.Rect(*rect), filename=img_path)

            elif ot == "add_watermark":
                text = op.get("text", "DRAFT")
                page_spec = op.get("pages", "all")
                font_size = float(op.get("font_size", 60))
                color = op.get("color", [0.8, 0.8, 0.8])
                if isinstance(color, list):
                    color = tuple(float(c) for c in color)
                rotation = float(op.get("rotation", 45))
                opacity = float(op.get("opacity", 0.3))

                text_len = fitz.get_text_length(
                    text, fontname="helv", fontsize=font_size)
                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    rect = page.rect
                    center = fitz.Point(rect.width / 2, rect.height / 2)
                    # Straight text centered on the page, then rotated as one
                    # unit around the center via morph. insert_text's own
                    # `rotate` only takes quarter turns — and combining it
                    # with a morph used to shear each glyph individually.
                    m = fitz.Matrix(1, 1)
                    m.prerotate(rotation)  # positive = the usual ↗ diagonal
                    page.insert_text(
                        fitz.Point(center.x - text_len / 2,
                                   center.y + font_size * 0.35),
                        text,
                        fontsize=font_size,
                        fontname="helv",
                        color=color,
                        overlay=True,
                        fill_opacity=opacity,
                        stroke_opacity=opacity,
                        morph=(center, m),
                    )

            elif ot == "replace_text":
                find = op.get("find", "")
                replace = op.get("replace", "")
                if not find:
                    errors.append(f"Op #{idx} replace_text: 'find' text required")
                    continue
                page_spec = op.get("pages", "all")
                case_sensitive = bool(op.get("case_sensitive", False))
                total, pages_hit = 0, 0

                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    hits = page.search_for(find)
                    if case_sensitive:
                        # search_for matches case-insensitively; keep only
                        # hits whose underlying text contains the exact form.
                        hits = [
                            r for r in hits
                            if find in page.get_text(
                                "text", clip=fitz.Rect(r) + (-1, -1, 1, 1))
                        ]
                    if not hits:
                        continue

                    # Capture style + free width BEFORE redacting (redaction
                    # removes the spans the probes read), then redact, then
                    # reinsert.
                    words = page.get_text("words")
                    plans = []
                    for r in hits:
                        size, color, fontname, baseline_y = _text_style_at(page, r)
                        # The replacement may run past the original footprint
                        # up to the next word on the same line (minus one
                        # space width) — or to the right margin when the hit
                        # ends its line.
                        next_x0 = min(
                            (w[0] for w in words
                             if w[1] < r.y1 and w[3] > r.y0 and w[0] >= r.x1 - 1),
                            default=page.rect.x1 - 36,
                        )
                        max_w = max(r.width, next_x0 - r.x0 - 0.25 * size)
                        plans.append((fitz.Rect(r), size, color, fontname,
                                      baseline_y, max_w))
                        page.add_redact_annot(r)
                    # IMAGE_NONE: never damage figures the hit rect overlaps.
                    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

                    if replace:
                        for r, size, color, fontname, baseline_y, max_w in plans:
                            fs = size
                            # Shrink until the replacement fits the free width
                            # so it never collides with its neighbors.
                            while fs > 5 and fitz.get_text_length(
                                    replace, fontname=fontname,
                                    fontsize=fs) > max_w:
                                fs -= 0.5
                            page.insert_text(
                                fitz.Point(r.x0, baseline_y), replace,
                                fontsize=fs, fontname=fontname, color=color,
                            )
                    total += len(hits)
                    pages_hit += 1

                if total == 0:
                    errors.append(f"Op #{idx} replace_text: '{find}' not found")
                else:
                    what = "deleted" if not replace else f"replaced with '{replace}'"
                    notes.append(
                        f"replace_text: {total} occurrence(s) of '{find}' "
                        f"{what} across {pages_hit} page(s)")

            elif ot == "redact":
                find = op.get("find")
                rect = op.get("rect")
                if not find and not (rect and len(rect) == 4):
                    errors.append(
                        f"Op #{idx} redact: needs 'find' text or a 4-number 'rect'")
                    continue
                page_spec = op.get("pages", "all")
                fill = op.get("fill", [0, 0, 0])
                if isinstance(fill, list) and len(fill) == 3:
                    fill = tuple(float(c) for c in fill)
                else:
                    fill = (0, 0, 0)
                count = 0

                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    targets = page.search_for(find) if find else [fitz.Rect(*rect)]
                    if not targets:
                        continue
                    for r in targets:
                        page.add_redact_annot(r, fill=fill)
                    # Privacy-grade: blank image pixels under the region too.
                    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_PIXELS)
                    count += len(targets)

                if count == 0:
                    errors.append(
                        f"Op #{idx} redact: nothing matched"
                        + (f" '{find}'" if find else ""))
                else:
                    notes.append(f"redact: {count} region(s) permanently removed")

            elif ot == "add_annotation":
                pi = int(op.get("page", 0))
                if pi >= doc.page_count:
                    errors.append(f"Op #{idx} add_annotation: page {pi} out of range")
                    continue
                page = doc[pi]
                annot_type = op.get("annotation_type", "highlight")
                rect = op.get("rect", [100, 700, 400, 720])
                r = fitz.Rect(*rect)
                color = op.get("color", [1, 1, 0])
                if isinstance(color, list):
                    color = tuple(float(c) for c in color)
                content = op.get("content", "")

                if annot_type == "highlight":
                    annot = page.add_highlight_annot(r)
                elif annot_type == "underline":
                    annot = page.add_underline_annot(r)
                elif annot_type == "strikeout":
                    annot = page.add_strikeout_annot(r)
                elif annot_type == "text_note":
                    annot = page.add_text_annot(fitz.Point(r.x0, r.y0), content)
                elif annot_type == "rectangle":
                    annot = page.add_rect_annot(r)
                else:
                    errors.append(f"Op #{idx} add_annotation: unknown type '{annot_type}'")
                    continue

                if annot and color and annot_type != "text_note":
                    annot.set_colors(stroke=color)
                    annot.update()
                if annot and content and annot_type != "text_note":
                    annot.set_info(content=content)
                    annot.update()

            # =============================================================
            # SECURITY & OPTIMIZATION
            # =============================================================

            elif ot == "encrypt":
                user_pw = op.get("user_password") or op.get("password", "")
                owner_pw = op.get("owner_password", user_pw)
                perms_raw = op.get("permissions", {})

                # Accept both dict {"print": true} and list ["print", "copy"]
                if isinstance(perms_raw, list):
                    perms = {p: True for p in perms_raw}
                elif isinstance(perms_raw, dict):
                    perms = perms_raw
                else:
                    perms = {}

                perm_flags = fitz.PDF_PERM_ACCESSIBILITY
                if perms.get("print", True):
                    perm_flags |= fitz.PDF_PERM_PRINT | fitz.PDF_PERM_PRINT_HQ
                if perms.get("copy", True):
                    perm_flags |= fitz.PDF_PERM_COPY
                if perms.get("annotate", True):
                    perm_flags |= fitz.PDF_PERM_ANNOTATE
                if perms.get("modify", False):
                    perm_flags |= fitz.PDF_PERM_MODIFY

                # Save encrypted — applied at save time, store params for later
                doc._encrypt_user_pw = user_pw
                doc._encrypt_owner_pw = owner_pw
                doc._encrypt_perms = perm_flags

            elif ot == "decrypt":
                password = op.get("password", "")
                if doc.is_encrypted:
                    if not doc.authenticate(password):
                        errors.append(f"Op #{idx} decrypt: incorrect password")

            elif ot == "compress":
                # Compression applied at save time — store params
                doc._compress_garbage = int(op.get("garbage", 4))
                doc._compress_deflate = op.get("deflate", True)
                doc._compress_deflate_images = op.get("deflate_images", True)
                doc._compress_deflate_fonts = op.get("deflate_fonts", True)

            elif ot == "set_metadata":
                new_meta = {}
                for key in ("title", "author", "subject", "keywords", "creator", "producer"):
                    if op.get(key):
                        new_meta[key] = op[key]
                if new_meta:
                    doc.set_metadata(new_meta)

            # =============================================================
            # EXTRACTION
            # =============================================================

            elif ot == "extract_images":
                page_spec = op.get("pages", "all")
                output_dir = op.get("output_dir")
                if not output_dir:
                    output_dir = str(Path(path).parent / (Path(path).stem + "_images"))
                out_dir = _checked_resolved(output_dir)
                Path(out_dir).mkdir(parents=True, exist_ok=True)

                extracted = []
                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    for img_idx, img in enumerate(page.get_images(full=True)):
                        xref = img[0]
                        try:
                            base_image = doc.extract_image(xref)
                            ext = base_image["ext"]
                            img_bytes = base_image["image"]
                            fname = f"page{pi + 1}_img{img_idx + 1}.{ext}"
                            img_path = Path(out_dir) / fname
                            img_path.write_bytes(img_bytes)
                            extracted.append(str(img_path))
                        except Exception as e:
                            logger.warning(f"Failed to extract image xref={xref}: {e}")

                if extracted:
                    errors.append(f"Op #{idx} extract_images: extracted {len(extracted)} images to {_to_agents_relative(out_dir)}")

            elif ot == "ocr":
                language = op.get("language", "eng")
                page_spec = op.get("pages", "all")
                output_path = op.get("output_path")
                requested_dpi = op.get("dpi")

                for pi in _parse_pages(page_spec, doc.page_count):
                    page = doc[pi]
                    try:
                        # Render the page and let Tesseract produce a 1-page
                        # PDF with an invisible text layer, then swap it in.
                        # (A TextPage from get_textpage_ocr() is extraction-
                        # only — it never modifies the document, so the saved
                        # PDF would stay unsearchable.) The raster must stay
                        # RGB — a csGRAY pixmap makes pdfocr_tobytes return
                        # an EMPTY text layer (pymupdf 1.27) — so output size
                        # is controlled by capping dpi at the scan's own
                        # resolution instead of upsampling to 300.
                        pix = page.get_pixmap(
                            dpi=_ocr_raster_dpi(page, requested_dpi),
                        )
                        ocr_pdf = fitz.open(
                            "pdf", pix.pdfocr_tobytes(language=language)
                        )
                        doc.delete_page(pi)
                        doc.insert_pdf(
                            ocr_pdf, from_page=0, to_page=0, start_at=pi
                        )
                        ocr_pdf.close()
                    except Exception as e:
                        errors.append(f"Op #{idx} ocr: page {pi} failed: {e}")

                if output_path:
                    out = _checked_resolved(output_path)
                    Path(out).parent.mkdir(parents=True, exist_ok=True)
                    doc.save(out)

            # =============================================================
            # UNKNOWN
            # =============================================================

            else:
                logger.warning(f"edit_pdf: unknown operation '{ot}', skipping")
                errors.append(f"Op #{idx}: unknown operation '{ot}'")

        except Exception as exc:
            errors.append(f"Op #{idx} {ot}: {exc}")
            logger.warning(f"edit_pdf op #{idx} '{ot}' failed: {exc}")

    # Save with encryption/compression if requested
    save_kwargs = {}

    # Encryption params
    if hasattr(doc, "_encrypt_user_pw"):
        save_kwargs["user_pw"] = doc._encrypt_user_pw
        save_kwargs["owner_pw"] = doc._encrypt_owner_pw
        save_kwargs["permissions"] = doc._encrypt_perms
        save_kwargs["encryption"] = fitz.PDF_ENCRYPT_AES_256

    # Compression params
    garbage = getattr(doc, "_compress_garbage", 0)
    if garbage:
        save_kwargs["garbage"] = garbage
    if getattr(doc, "_compress_deflate", False):
        save_kwargs["deflate"] = True
    if getattr(doc, "_compress_deflate_images", False):
        save_kwargs["deflate_images"] = True
    if getattr(doc, "_compress_deflate_fonts", False):
        save_kwargs["deflate_fonts"] = True

    # Save — pymupdf can't overwrite the source file directly, so save to a
    # deterministic temp then replace (deterministic so the PARENT can clean
    # the orphan after a worker kill — see _WORKER_TMP_SUFFIX).
    size_before = os.path.getsize(path)
    tmp_path = path + _WORKER_TMP_SUFFIX
    try:
        doc.save(tmp_path, **save_kwargs)
        doc.close()
        os.replace(tmp_path, path)
    except Exception:
        doc.close()
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise
    size_after = os.path.getsize(path)

    msg = f"PDF saved: {_to_agents_relative(path)} ({len(ops)} operations applied)"
    msg += _dropped_note(dropped)
    if garbage:
        pct = (1 - size_after / max(size_before, 1)) * 100
        msg += f"\nCompression: {size_before / 1024:.0f}KB → {size_after / 1024:.0f}KB ({pct:.0f}% reduction)"
    if notes:
        msg += "\n" + "\n".join(f"  - {n}" for n in notes)
    if errors:
        msg += f"\n\nWarnings/Errors ({len(errors)}):\n" + "\n".join(f"  - {e}" for e in errors)
    return msg


# ---------------------------------------------------------------------------
# PDF to Images
# ---------------------------------------------------------------------------


_RENDER_ADVICE = "Render fewer pages per call (`pages`) or lower `dpi`"


def _pdf_to_images_core(path: str, tmp_dir: str, page_spec, dpi: int,
                        fmt: str) -> list[dict]:
    """Worker core: render every requested page into the parent-named temp
    dir. The parent moves the COMPLETE set into the real output dir only on
    success — a deadline kill must not leave a partial page set that looks
    complete on `ls`. Unbounded pixmaps (degenerate page geometry × dpi) are
    exactly what the worker's RLIMIT_AS is for."""
    import fitz

    if os.path.isdir(tmp_dir):
        shutil.rmtree(tmp_dir)
    os.makedirs(tmp_dir)

    doc = fitz.open(path)
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    saved = []
    for pi in _parse_pages(page_spec, doc.page_count):
        page = doc[pi]
        pix = page.get_pixmap(matrix=mat)
        fname = f"page_{pi + 1:03d}.{fmt}"
        out_path = os.path.join(tmp_dir, fname)
        if fmt in ("jpg", "jpeg"):
            pix.save(out_path, jpg_quality=95)
        else:
            pix.save(out_path)
        saved.append({"name": fname, "width": pix.width, "height": pix.height})
    doc.close()
    return saved


async def handle_pdf_to_images(args: dict) -> str:
    path = await _resolve_path(args["path"])
    if not Path(path).exists():
        return f"Error: File not found: {args['path']}"

    output_dir = args.get("output_dir")
    if not output_dir:
        output_dir = str(Path(path).parent / (Path(path).stem + "_pages"))
    out_dir = await _resolve_path(output_dir, writing=True)

    page_spec = args.get("pages", "all")
    dpi = int(args.get("dpi", 150))
    fmt = args.get("format", "png").lower()

    tmp_dir = out_dir.rstrip("/") + _WORKER_TMP_SUFFIX
    try:
        saved = await run_parse(
            _pdf_to_images_core, path, tmp_dir, page_spec, dpi, fmt,
            _advice=_RENDER_ADVICE,
        )
    except Exception:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise

    Path(out_dir).mkdir(parents=True, exist_ok=True)
    for s in saved:
        os.replace(os.path.join(tmp_dir, s["name"]), os.path.join(out_dir, s["name"]))
    shutil.rmtree(tmp_dir, ignore_errors=True)

    # Push first page inline for preview
    if saved:
        img_bytes = (Path(out_dir) / saved[0]["name"]).read_bytes()
        mime = "image/png" if fmt == "png" else "image/jpeg"
        await _push_image_preview(img_bytes, mime, f"Page 1 of {Path(path).name}")

    return (
        f"Rendered {len(saved)} pages from {_to_agents_relative(path)} at {dpi}dpi.\n"
        f"Output: {_to_agents_relative(out_dir)}/\n"
        f"Files: {', '.join(s['name'] for s in saved[:5])}"
        + (f" ... +{len(saved) - 5} more" if len(saved) > 5 else "")
    )


# ---------------------------------------------------------------------------
# Screenshot Document (visual feedback for LLM)
# ---------------------------------------------------------------------------

_SCREENSHOT_EXTS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
                    ".odt", ".ods", ".odp", ".csv", ".rtf", ".html"}
_MAX_INLINE_PAGES = 10
# Payload caps (see handle_screenshot_document): long-edge px per page
# (dpi-aware — the high cap keeps A4@300 ≈ 3508 px untouched, the skill's
# documented handwriting remedy), PNG→JPEG size fallback, and the per-call
# total across all pages (truncates with a note, never silently).
_SCREENSHOT_LONG_EDGE_STD = 2000
_SCREENSHOT_LONG_EDGE_HIGH = 3600
_SCREENSHOT_PNG_FALLBACK_BYTES = int(1.5 * 1024 * 1024)
_SCREENSHOT_TOTAL_BYTES = 8 * 1024 * 1024


def _render_screenshot_core(pdf_path: str, pages_spec, dpi: int) -> dict:
    """Worker core: the fitz render loop with the payload caps below.
    Returns raw bytes — base64 and mcp.types stay in the parent (they never
    cross the pipe).

    Vision-sized payloads: the 300 s timeouts this tool hit were the
    RESPONSE path for multi-MB base64, never rendering (the same page
    renders in under a second) — so cap pixels and bytes here. The long-edge
    cap is dpi-aware: a flat cap would bite exactly on the explicit dpi:300
    calls the skill mandates for handwriting (A4@300 ≈ 3508 px passes
    untouched). Encoding follows CONTENT: pages with a text layer stay PNG
    (JPEG ringing is worst on crisp text) with a byte-size JPEG fallback;
    scanned pages go straight to JPEG q85. A total budget truncates the
    batch with a note."""
    import fitz

    zoom = dpi / 72.0
    doc = fitz.open(pdf_path)
    total = len(doc)
    page_indices = _parse_pages(pages_spec, total)
    if not page_indices:
        doc.close()
        return {"total": total, "images": [], "rendered": [],
                "capped": False, "truncated_at": None}

    capped = False
    if len(page_indices) > _MAX_INLINE_PAGES:
        page_indices = page_indices[:_MAX_INLINE_PAGES]
        capped = True

    long_edge_cap = (
        _SCREENSHOT_LONG_EDGE_STD if dpi <= 200 else _SCREENSHOT_LONG_EDGE_HIGH
    )
    images: list = []
    rendered: list = []
    total_bytes = 0
    truncated_at = None
    for pi in page_indices:
        page = doc[pi]
        long_edge = max(page.rect.width, page.rect.height) * zoom
        eff_zoom = zoom if long_edge <= long_edge_cap else (
            zoom * (long_edge_cap / long_edge)
        )
        pix = page.get_pixmap(matrix=fitz.Matrix(eff_zoom, eff_zoom))
        scanned = not page.get_text().strip()
        if scanned:
            img_bytes = pix.tobytes("jpg", jpg_quality=85)
            mime = "image/jpeg"
        else:
            img_bytes = pix.tobytes("png")
            mime = "image/png"
            if len(img_bytes) > _SCREENSHOT_PNG_FALLBACK_BYTES:
                img_bytes = pix.tobytes("jpg", jpg_quality=85)
                mime = "image/jpeg"
        if images and total_bytes + len(img_bytes) > _SCREENSHOT_TOTAL_BYTES:
            truncated_at = pi + 1
            break
        total_bytes += len(img_bytes)
        images.append({"data": img_bytes, "mime": mime})
        rendered.append({
            "page": pi + 1,
            "width_in": round(page.rect.width / 72, 2),
            "height_in": round(page.rect.height / 72, 2),
        })

    doc.close()
    return {"total": total, "images": images, "rendered": rendered,
            "capped": capped, "truncated_at": truncated_at}


async def handle_screenshot_document(args: dict) -> list:
    """Render document pages as ImageContent for LLM visual inspection.

    Returns list of ImageContent + TextContent (not pushed to dashboard).
    """
    import base64
    import tempfile

    from mcp.types import ImageContent, TextContent

    path = await _resolve_path(args.get("path", ""))
    if not os.path.isfile(path):
        return [TextContent(type="text", text=f"Error: file not found: {_to_agents_relative(path)}")]

    ext = Path(path).suffix.lower()
    if ext not in _SCREENSHOT_EXTS:
        return [TextContent(type="text",
                text=f"Error: unsupported format '{ext}'. Supported: PDF, DOCX, XLSX, PPTX and other Office formats.")]

    pages_spec = args.get("pages", "1")
    dpi = int(args.get("dpi", 150))
    sheet = args.get("sheet")

    pdf_path = path
    temp_dir = None

    try:
        # For non-PDF: convert to temp PDF via LibreOffice
        if ext != ".pdf":
            temp_dir = tempfile.mkdtemp(dir=str(Path(path).parent))
            convert_path = path

            # Excel: set fit-to-width page setup + handle sheet selection.
            # The prep full-DOM-loads the workbook (bomb risk) — worker
            # child. Sheet names ride back with it so no un-isolated
            # openpyxl load remains in the parent (even a read_only load
            # eagerly parses the whole sharedStrings table).
            if ext in (".xlsx", ".xls"):
                try:
                    convert_path, sheetnames = await run_parse(
                        _excel_prepare_for_screenshot, path, temp_dir,
                        _advice=_RENDER_ADVICE,
                    )
                    if sheet is not None:
                        # Sheet selection: override pages to target sheet
                        pages_spec = str(
                            _sheet_index_from_names(sheetnames, sheet) + 1
                        )
                except Exception as e:
                    logger.warning(f"screenshot excel prep failed: {e}, using original")
                    convert_path = path
            elif sheet is not None and ext in (".ods", ".csv"):
                # ODS/CSV: openpyxl can't open either format, so the old
                # name lookup ALWAYS failed here — numeric sheet selection
                # is honored, names warn (the reachable subset of the old
                # behavior, minus the pointless workbook load).
                if isinstance(sheet, int) or str(sheet).isdigit():
                    pages_spec = str(int(sheet) + 1)
                else:
                    logger.warning(
                        f"screenshot sheet lookup by name unsupported for {ext}"
                    )

            pdf_path = await _libreoffice_convert(convert_path, "pdf", temp_dir)

        # Render in a bounded worker child (fitz pixmaps are the memory
        # hazard); the parent only b64-encodes and assembles content items.
        render = await run_parse(
            _render_screenshot_core, pdf_path, pages_spec, dpi,
            _advice=_RENDER_ADVICE,
        )

        if not render["rendered"]:
            return [TextContent(
                type="text",
                text=f"No valid pages to render (document has {render['total']} pages).",
            )]

        content_items: list = []
        for img in render["images"]:
            content_items.append(ImageContent(
                type="image",
                data=base64.b64encode(img["data"]).decode(),
                mimeType=img["mime"],
            ))

        # Append text summary after images
        fname = Path(path).name
        lines = [f"Rendered {len(render['rendered'])} page(s) from {fname} at {dpi} DPI:"]
        for r in render["rendered"]:
            lines.append(f"  Page {r['page']}: {r['width_in']} x {r['height_in']} inches")
        if render["capped"]:
            lines.append(f"  (capped at {_MAX_INLINE_PAGES} pages — use specific page ranges for more)")
        if render["truncated_at"] is not None:
            lines.append(
                f"  (stopped before page {render['truncated_at']}: inline image budget "
                "reached — request the remaining pages in a follow-up call)"
            )
        content_items.append(TextContent(type="text", text="\n".join(lines)))
        return content_items

    except Exception as e:
        return [TextContent(type="text", text=f"Error rendering {_to_agents_relative(path)}: {e}")]
    finally:
        if temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


def _sheet_index_from_names(names: list, sheet) -> int:
    """Get 0-based sheet index from name or numeric index (parent-side —
    the names come back from the prep worker, never from an openpyxl load
    in the server process)."""
    if isinstance(sheet, int) or (isinstance(sheet, str) and sheet.isdigit()):
        return int(sheet)
    return names.index(str(sheet)) if str(sheet) in names else 0


def _excel_prepare_for_screenshot(path: str, temp_dir: str) -> tuple:
    """Worker core: create a temp copy of an Excel file with fit-to-width
    page setup. Returns (temp_path, sheetnames) — the full-DOM load below is
    exactly the allocation profile that took the read path down, so this
    always runs in a bounded child.

    Safeguards based on column count:
      <= 10 columns: portrait, fit to 1 page wide
      11-20 columns: landscape, fit to 1 page wide
      > 20 columns:  landscape, no fit (would be unreadably small)
    Rows always flow naturally across pages.
    """
    import openpyxl
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        # Count used columns
        max_col = ws.max_column or 1

        if max_col <= 10:
            # Portrait, fit all columns to 1 page wide
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        elif max_col <= 20:
            # Landscape, fit all columns to 1 page wide
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            # Too many columns — landscape but let it paginate
            ws.page_setup.orientation = "landscape"

    temp_path = os.path.join(temp_dir, Path(path).name)
    names = list(wb.sheetnames)
    wb.save(temp_path)
    wb.close()
    return temp_path, names


# ---------------------------------------------------------------------------
# Images to PDF
# ---------------------------------------------------------------------------


def _images_to_pdf_core(image_paths: list, out: str, page_size: str,
                        fit: str) -> int:
    """Worker core: build the PDF from pre-resolved, existing image paths
    and save atomically. A bomb image (fitz decodes it for dimensions and
    embedding) dies at the child's limit, not the server's."""
    import fitz

    # Page dimensions in points
    sizes = {
        "a4": (595, 842),
        "letter": (612, 792),
        "a3": (842, 1190),
    }

    doc = fitz.open()

    for img_path in image_paths:
        # Get image dimensions
        img_doc = fitz.open(img_path)
        if img_doc.page_count == 0:
            img_doc.close()
            continue
        img_rect = img_doc[0].rect
        img_w, img_h = img_rect.width, img_rect.height
        img_doc.close()

        if page_size == "original":
            pw, ph = img_w, img_h
        else:
            pw, ph = sizes.get(page_size, (595, 842))
            # Auto-rotate page to match image orientation
            if (img_w > img_h and pw < ph) or (img_h > img_w and ph < pw):
                pw, ph = ph, pw

        page = doc.new_page(width=pw, height=ph)

        if fit == "stretch":
            img_rect = fitz.Rect(0, 0, pw, ph)
        elif fit == "cover":
            scale = max(pw / img_w, ph / img_h)
            new_w, new_h = img_w * scale, img_h * scale
            x = (pw - new_w) / 2
            y = (ph - new_h) / 2
            img_rect = fitz.Rect(x, y, x + new_w, y + new_h)
        else:  # contain
            scale = min(pw / img_w, ph / img_h)
            new_w, new_h = img_w * scale, img_h * scale
            x = (pw - new_w) / 2
            y = (ph - new_h) / 2
            img_rect = fitz.Rect(x, y, x + new_w, y + new_h)

        page.insert_image(img_rect, filename=img_path)

    inserted = doc.page_count
    tmp = out + _WORKER_TMP_SUFFIX
    doc.save(tmp)
    doc.close()
    os.replace(tmp, out)
    return inserted


async def handle_images_to_pdf(args: dict) -> str:
    images = args.get("images", [])
    output_path = args.get("output_path")
    if not output_path:
        return "Error: output_path is required"
    if not images:
        return "Error: images list is empty"

    out = await _resolve_path(output_path, writing=True)
    Path(out).parent.mkdir(parents=True, exist_ok=True)

    page_size = args.get("page_size", "a4").lower()
    fit = args.get("fit", "contain")

    # Pre-resolve in the parent; a missing image skips with a warning (the
    # old per-image containment), it never fails the call.
    image_paths = []
    for img_input in images:
        try:
            img_path = await _resolve_path(img_input)
        except Exception as exc:
            logger.warning(f"images_to_pdf: skipping {img_input} ({exc})")
            continue
        if not Path(img_path).exists():
            logger.warning(f"images_to_pdf: skipping {img_input} (not found)")
            continue
        image_paths.append(img_path)

    try:
        await run_parse(
            _images_to_pdf_core, image_paths, out, page_size, fit,
            _advice=_RENDER_ADVICE,
        )
    except Exception:
        with contextlib.suppress(OSError):
            os.unlink(out + _WORKER_TMP_SUFFIX)
        raise

    await _push_preview(out)
    return f"PDF created: {_to_agents_relative(out)} ({len(images)} images, {page_size})"


# ---------------------------------------------------------------------------
# Convert document
# ---------------------------------------------------------------------------


async def handle_convert_document(args: dict) -> str:
    input_path = await _resolve_path(args["input_path"])
    if not Path(input_path).exists():
        return f"Error: File not found: {args['input_path']}"

    output_format = args.get("output_format", "pdf")
    output_path = args.get("output_path")

    if output_path:
        output_dir = str(Path(await _resolve_path(output_path, writing=True)).parent)
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    else:
        output_dir = str(Path(input_path).parent)

    ext = Path(input_path).suffix.lower()

    # Markdown → PDF via WeasyPrint
    if ext in (".md", ".markdown") and output_format == "pdf":
        content = Path(input_path).read_text(encoding="utf-8")
        out = str(Path(output_dir) / (Path(input_path).stem + ".pdf"))
        # Internal handler calls pass CONTAINER-ABSOLUTE paths — the display
        # form does not survive a _resolve_path round-trip (see pdf_to_images).
        await handle_write_pdf({
            "path": out,
            "content": content,
            "content_type": "markdown",
        })
        return f"Converted: {_to_agents_relative(out)}"

    # Image → PDF via pymupdf
    if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".tif", ".webp") and output_format == "pdf":
        out = str(Path(output_dir) / (Path(input_path).stem + ".pdf"))
        await handle_images_to_pdf({
            "images": [input_path],
            "output_path": out,
            "page_size": "original",
        })
        return f"Converted: {_to_agents_relative(out)}"

    # PDF → PNG via pymupdf
    if ext == ".pdf" and output_format in ("png", "jpg", "jpeg"):
        out_dir = str(Path(output_dir) / Path(input_path).stem)
        await handle_pdf_to_images({
            "path": input_path,
            "output_dir": out_dir,
            "format": output_format,
            "dpi": 150,
        })
        return f"Converted: {_to_agents_relative(out_dir)}/"

    # LibreOffice headless for everything else. Deliberately NOT a worker
    # child: soffice is already a separate OS process, so the container cap
    # bounds it and a cgroup OOM kill targets its RSS, not the server;
    # _libreoffice_lock serializes runs and the 120s timeout kills hangs.
    # A worker wrapper could only strip the child's RLIMIT_AS (it survives
    # fork+exec — see isolation.py) or cap soffice confusingly.
    try:
        out = await _libreoffice_convert(input_path, output_format, output_dir)
    except RuntimeError as exc:
        return f"Conversion error: {exc}"

    if output_path:
        final = await _resolve_path(output_path, writing=True)
        Path(out).rename(final)
        out = final

    await _push_preview(out)
    return f"Converted: {_to_agents_relative(out)}"
